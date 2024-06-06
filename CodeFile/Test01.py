#改进作业：
#将生成的excel里第一个表格里面的数据复制到第二个表格里面。

# 编程完成以下任务:
# (1)创建一个 generateTxt(txtName)函数，函数的功能为:
# 在程序源码同目录下创建一个名为 txtName 的文本文件
# 列和列之间使用英文逗号(“,”)隔开，
# 第一列“Co1”的数据为 1-10 之间的随机整数，
# 第二列为 10-20 之间的随机整数，
# 第三列为 20-30 之间的随机整数，共写入 10 行。
# (2)创建 changeTxtToExcel(txtName)函数，函数功能为:
# 将 tName 中的数据保存到相同名字的 excel 文件 (.xlsx) 中，
# 如上图所示要求:使用 Python 第三方库 openpyxl 创建 excel 文件
# (3) 在文件主函数内先后调用这两个函数，完成创建一个名为“Twotask.txt文件，
# 写入数据，然后转换成 Twotask.xlsx 文件。

from openpyxl import Workbook
from openpyxl import load_workbook
from random import randint

#创建函数用来生成每行随机数
def genarateTxt(txtName):
    # txtName = txtName + ".txt" #在名称后面加上后缀变为文本文档（注意：文件名包括文件后缀名）
    with open(txtName,'w') as fp: #打开原文件
        fp.write("Col1,Col2,Col3" + "\n") #写入第一行
        #生成随机数
        for i in range(0,10): #生成10行数据（range范围：左闭右开）
            c1 = str(randint(1,10)) #第一列生成1-10的随机数
            c2 = str(randint(10,20)) #第二列生成10-20的随机数
            c3 = str(randint(20,30)) #第三列生成20-30的随机数
            row = c1+","+c2+","+c3+"\n" #每一行的数据中间用逗号隔开，并且换行
            fp.write(row) #将生成的数据写入文件
    fp.close() #写入后关闭，释放资源

#创建函数将文本文档变为Excel格式文档
def changeTxtToEXCEL(txtName):
    wb = Workbook() #创建一个excel文件
    ws = wb.worksheets[0] #打开第一个工作表格
    Excel_name = txtName[:-3] + "xlsx" #将原文件的后缀去掉在名称后面加上文件后缀变为excel表格文档
    with open(txtName,"r") as fp: #打开原文本文档
        for line in fp: #逐行读取
            print(line) #读取出来
            line = line.strip().split(",") #去除首尾空格并以逗号为分割提取每列数据
            ws.append(line) #将本行提取出来的数据写入文件中
    wb.save(Excel_name) #保存excel文档,名称为“Twotask.xlsx”

#改进部分
#新建一个工作表，然后将sheet1的数据复制到sheet2中
#新建函数
def copyEXCEL(txtName):
    Excel_name = txtName[:-3] + "xlsx"
    #首先需要打开该工作簿
    wb = load_workbook(Excel_name)
    #将sheet的内容复制
    ws_1 = wb.worksheets[0] #获取工作簿的第一个工作表
    ws_2 = wb.copy_worksheet(ws_1) #调用openpyxl的copy_worksheet()方法进行复制
    #复制完毕
    wb.save(Excel_name) #保存文件

def main():
    fn = "Twotask.txt" #原文件名称
    genarateTxt(fn) #调用函数写入数据
    changeTxtToEXCEL(fn) #调用函数将数据保存到表格中
    copyEXCEL(fn) #复制工作表的内容

if __name__ == "__main__": #主方法执行
    main()

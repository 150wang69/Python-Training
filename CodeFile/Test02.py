# 将excel文件“每个人的爱好.xlsx”复制到代码目录下，并完成以下任务
# （1）使用openpyxl包操作打开此文件
# （2）对每个人的爱好进行汇总，并将其增加在最后一列.

from openpyxl import load_workbook

def sumEXCEl(FileName):
    wb = load_workbook(FileName) #打开使用文件
    #对每个人的爱好进行汇总
    ws = wb.worksheets[0] #打开该文件的第一个工作表（指定工作表）
    rowNum = ws.max_row #返回最大行
    colNum = ws.max_column #返回最大列
    # print(rowNum,colNum)

    ws.cell(row=1,column=colNum+1,value="所有爱好") #单元格写入数据
    #汇总每个人的爱好
    for i in range(2,rowNum+1):
        all_haihao = ""
        for j in range(2,colNum+1):
            if(ws.cell(row=i,column=j).value) == "是": #判断值为“是”，则获取row=1，column=j的值
                cell_txt = ws.cell(row=1, column=j).value  # 获取目标单元格的值
                all_haihao = all_haihao + cell_txt +" " #获取后加到all-haihao变量上,中间用空格隔开
        print(ws.cell(row=i,column=1).value + "的所有爱好是" + all_haihao) #进行汇总打印
        ws.cell(row=i,column=colNum+1).value = all_haihao #汇总后添加到最后一列
        wb.save("每个人的爱好汇总.xlsx") #另存

def main():
    fn = "每个人的爱好.xlsx"
    sumEXCEl(fn)

if __name__ == '__main__':
    main()

"""
（1）使用requests请求链家二手房网站，自行选择一个城市数据进行爬取
参考URL：https://cq.lianjia.com/ershoufang/
（2）使用etree解析html，获取二手房的信息，如下图所示，至少获取红
框中标注的7条信息
"""
import requests
from openpyxl import Workbook

#首先使用requests对网站进行url的爬取
from lxml import etree

wb = Workbook() #创建工作薄
ws = wb.active #由于在创建工作簿的同时也创建一个工作表，所以使用active直接调用该工作表

ws['A1'] = "名称"
ws['B1'] = "所处小区"
ws['C1'] = "楼房"
ws['D1'] = "房间信息"
ws['E1'] = "关注度"

response = requests.get("https://cq.lianjia.com/ershoufang/")#发送请求
response.encoding="utf-8"
print(response)#接受响应
# print(response.text)#获取url的源代码
html = etree.HTML(response.text)

#通过xpath筛选出来所需要的代码信息

#获取房间的信息
houseInfos = html.xpath(r"//div[@class='houseInfo']/text()")
for i in range(len(houseInfos)):
    # print(houseInfos[i])
    ws['D{}'.format(i + 2)] = houseInfos[i]

#获取房子的名称
houseNames = html.xpath(r'//div[@class="title"]/a/text()')
for i in range(len(houseNames)):
    ws['A{}'.format(i + 2)] = houseNames[i]

#获取房间的小区
houseLoc_roads = html.xpath(r"//div[@class='positionInfo']/a[position()>1]/text()")
for i in range(len(houseLoc_roads)):
    ws['B{}'.format(i + 2)] = houseLoc_roads[i]

#获取房间的楼房
houseLoc_regions = html.xpath(r"//div[@class='positionInfo']/a[@data-el='region']/text()")
for i in range(len(houseLoc_regions)):
    ws['C{}'.format(i + 2)] = houseLoc_regions[i]

#将小区和楼房匹配
# for i in range(len(houseLoc_roads)):
#     coLocInfo = (houseLoc_regions[i],houseLoc_roads[i])
#     print(coLocInfo)
#     ws['B{}'.format(i + 2)] = coLocInfo

#获取关注度
houseFollows = html.xpath(r"//div[@class='followInfo']/text()")
for i in range(len(houseFollows)):
    ws['E{}'.format(i + 2)] = houseFollows[i]

#保存数据并关闭
wb.save("房子信息统计.xlsx")
wb.close()
